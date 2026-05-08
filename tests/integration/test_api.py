from io import BytesIO
from pathlib import Path

import pytest
from market.models import ConfidenceStatus, Listing, RawListing, ReviewStatus, SearchRun
from openpyxl import load_workbook

from packages.quality_gates.data_visibility_gate import run_data_visibility_gate
from packages.scrapers.dfimoveis import DFImoveisOfflineScraper
from packages.scrapers.persistence import persist_scraper_result


def test_health_endpoint(client):
    response = client.get("/api/v1/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"


@pytest.mark.django_db
def test_demo_endpoint_creates_data(client):
    response = client.post("/api/v1/search-runs/demo")
    assert response.status_code == 200
    payload = response.json()
    assert payload["total_raw_collected"] == 6
    assert payload["total_reliable"] == 2


def test_summary_endpoint_reports_raw_universe(client, demo_run):
    response = client.get("/api/v1/market/summary")
    assert response.status_code == 200
    payload = response.json()
    assert payload["total_raw_collected"] == demo_run.total_raw_collected
    assert payload["total_raw_without_listing"] == 0
    assert payload["total_opportunities"] == 2


def test_dashboard_opens(client, demo_run):
    response = client.get("/dashboard/")
    content = response.content.decode()
    assert response.status_code == 200
    assert "Jarjour Market Intelligence" in content
    assert "Leitura executiva do mercado imobiliário" in content
    assert "Brutos sem normalização" in content
    assert "Imóveis normalizados" in content
    assert "Exportar Excel Carol" in content
    for card in [
        "Anúncios coletados",
        "Normalizados",
        "Brutos sem normalização",
        "Confiáveis",
        "Para revisar",
        "Incompletos",
        "Oportunidades",
    ]:
        assert card in content
    for action in [
        "Aprovar",
        "Marcar para revisão",
        "Rejeitar",
        "Marcar como oportunidade",
    ]:
        assert action in content


@pytest.mark.django_db
def test_dashboard_and_gate_show_raw_without_listing(client):
    run = SearchRun.objects.create(source_name="critical")
    raw_with_listing = RawListing.objects.create(
        search_run=run,
        source_name="critical",
        raw_title="Sala normalizada",
    )
    RawListing.objects.create(search_run=run, source_name="critical", raw_title="Bruto pendente 1")
    RawListing.objects.create(search_run=run, source_name="critical", raw_title="Bruto pendente 2")
    Listing.objects.create(
        raw_listing=raw_with_listing,
        source_name="critical",
        endereco="SCS Quadra 2",
        bairro="Asa Sul",
        tipo_imovel="Sala",
        finalidade="Aluguel",
        confidence_status=ConfidenceStatus.RELIABLE,
    )

    response = client.get("/api/v1/market/summary")
    payload = response.json()
    assert payload["total_raw_collected"] == 3
    assert payload["total_normalized"] == 1
    assert payload["total_raw_without_listing"] == 2

    dashboard = client.get("/dashboard/")
    content = dashboard.content.decode()
    assert dashboard.status_code == 200
    assert "Brutos sem normalização" in content
    assert "Existem 2 anúncios brutos ainda sem normalização" in content
    assert "Eles estão preservados na base e não foram descartados." in content

    result = run_data_visibility_gate()
    assert result.passed, result.message


def test_excel_export_endpoint_has_workbook(client, demo_run):
    response = client.get("/api/v1/exports/carol-xlsx")
    assert response.status_code == 200
    wb = load_workbook(filename=BytesIO(response.content), read_only=True)
    assert "Pesquisa Carol" in wb.sheetnames


def test_review_listing_endpoint_updates_listing(client, demo_run):
    listing = Listing.objects.filter(raw_listing__search_run=demo_run).first()

    response = client.post(
        f"/api/v1/listings/{listing.id}/review",
        data={
            "decision": ReviewStatus.OPPORTUNITY,
            "comment": "Boa oportunidade",
            "reviewed_by": "Carol",
        },
        content_type="application/json",
    )

    listing.refresh_from_db()
    demo_run.refresh_from_db()
    payload = response.json()
    assert response.status_code == 200
    assert payload["id"] == listing.id
    assert payload["review_status"] == ReviewStatus.OPPORTUNITY
    assert payload["is_opportunity"] is True
    assert payload["last_review"]["decision"] == ReviewStatus.OPPORTUNITY
    assert payload["message"] == "Revisão registrada com rastreabilidade."
    assert demo_run.total_opportunities >= 1


def test_review_listing_endpoint_returns_404_for_missing_listing(client, db):
    response = client.post(
        "/api/v1/listings/999999/review",
        data={"decision": ReviewStatus.APPROVED},
        content_type="application/json",
    )

    assert response.status_code == 404


def test_review_listing_endpoint_rejects_invalid_decision(client, demo_run):
    listing = Listing.objects.filter(raw_listing__search_run=demo_run).first()

    response = client.post(
        f"/api/v1/listings/{listing.id}/review",
        data={"decision": "INVALID"},
        content_type="application/json",
    )

    assert response.status_code == 400
    assert "Decision invalid" in response.json()["detail"]


def test_dashboard_review_action_rejects_without_reducing_raw_total(client, demo_run):
    listing = Listing.objects.filter(raw_listing__search_run=demo_run).first()
    raw_total_before = demo_run.total_raw_collected

    response = client.post(
        f"/dashboard/listings/{listing.id}/review/",
        data={"decision": ReviewStatus.REJECTED, "reviewed_by": "Carol"},
    )

    listing.refresh_from_db()
    demo_run.refresh_from_db()
    assert response.status_code == 302
    assert listing.review_status == ReviewStatus.REJECTED
    assert Listing.objects.filter(id=listing.id).exists()
    assert RawListing.objects.filter(id=listing.raw_listing_id).exists()
    assert demo_run.total_raw_collected == raw_total_before


@pytest.mark.django_db
def test_normalization_endpoint_runs_for_search_run(client):
    run = SearchRun.objects.create(
        source_name="normalization-api",
        bairro="Asa Sul",
        tipo_imovel="Sala",
        finalidade="Aluguel",
    )
    RawListing.objects.create(
        search_run=run,
        source_name="normalization-api",
        source_url="https://fixture.local/api/one",
        raw_title="Sala comercial na Asa Sul",
        raw_address="CLS 308 Bloco B",
        raw_price="R$ 3.500,00",
        raw_area="45 m²",
    )

    response = client.post(
        "/api/v1/normalization/run",
        data={"search_run_id": run.id},
        content_type="application/json",
    )

    run.refresh_from_db()
    payload = response.json()
    assert response.status_code == 200
    assert payload["raw_processed"] == 1
    assert payload["listings_created"] == 1
    assert payload["message"] == "Normalizacao concluida sem apagar dados brutos."
    assert run.total_raw_collected == 1
    assert run.total_normalized == 1
    assert Listing.objects.filter(raw_listing__search_run=run, is_opportunity=True).count() == 0


@pytest.mark.django_db
def test_dashboard_and_excel_after_offline_normalization(client):
    fixture = Path("packages/scrapers/fixtures/dfimoveis_search_results_sample.html")
    html = fixture.read_text(encoding="utf-8")
    result = DFImoveisOfflineScraper().collect_from_html(html, evidence_path="test")
    run = persist_scraper_result(result, bairro="", tipo_imovel="", finalidade="Aluguel")

    client.post(
        "/api/v1/normalization/run",
        data={"search_run_id": run.id},
        content_type="application/json",
    )

    dashboard = client.get("/dashboard/")
    content = dashboard.content.decode()
    assert dashboard.status_code == 200
    assert "CLS 308 Bloco B" in content
    assert "Imóveis normalizados" in content or "ImÃ³veis normalizados" in content

    response = client.get("/api/v1/exports/carol-xlsx")
    wb = load_workbook(filename=BytesIO(response.content), read_only=True)
    headers = [cell.value for cell in next(wb["Pesquisa Carol"].iter_rows(max_row=1))]
    assert "ENDEREÇO" in headers or "ENDEREÃ‡O" in headers
