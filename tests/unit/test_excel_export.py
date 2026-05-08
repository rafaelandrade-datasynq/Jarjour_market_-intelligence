from io import BytesIO

from exports.services import (
    CAROL_HEADERS,
    CAROL_SHEET_NAMES,
    LINK_HEADERS,
    build_carol_workbook,
)
from market.models import RawListing, SearchRun
from openpyxl import load_workbook


def _sheet_rows(ws):
    return list(ws.iter_rows(min_row=2, values_only=True))


def _summary_value(workbook, label):
    ws = workbook["Resumo da Coleta"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] == label:
            return row[1]
    raise AssertionError(f"Summary label not found: {label}")


def test_carol_workbook_has_required_sheets_and_headers(demo_run):
    workbook = build_carol_workbook()

    assert workbook.sheetnames == CAROL_SHEET_NAMES
    assert [cell.value for cell in workbook["Pesquisa Carol"][1]] == CAROL_HEADERS
    assert [cell.value for cell in workbook["Links"][1]] == LINK_HEADERS


def test_pesquisa_carol_contains_non_discarded_listings(demo_run):
    workbook = build_carol_workbook()
    statuses = [row[12] for row in _sheet_rows(workbook["Pesquisa Carol"])]

    assert len(statuses) == 5
    assert "Descartado" not in statuses


def test_oportunidades_contains_only_opportunities(demo_run):
    workbook = build_carol_workbook()
    rows = _sheet_rows(workbook["Oportunidades"])

    assert len(rows) == 2
    assert {row[0] for row in rows} == {"CLS 308 Bloco B", "CLNW 10/11 Bloco C"}


def test_revisar_contains_only_review_needed_records(demo_run):
    workbook = build_carol_workbook()
    rows = _sheet_rows(workbook["Revisar"])
    statuses = {row[12] for row in rows}

    assert len(rows) == 3
    assert statuses == {"Para revisar", "Incompleto", "Duplicado provavel"}


def test_links_sheet_contains_clickable_urls(demo_run):
    workbook = build_carol_workbook()
    links = workbook["Links"]

    assert links.max_row == 7
    for row in range(2, links.max_row + 1):
        assert links.cell(row=row, column=4).value.startswith("https://example.local/")
        assert links.cell(row=row, column=4).hyperlink.target == links.cell(row=row, column=4).value


def test_resumo_da_coleta_contains_required_metrics(demo_run):
    workbook = build_carol_workbook()

    assert _summary_value(workbook, "Total bruto coletado") == 6
    assert _summary_value(workbook, "Total normalizado") == 6
    assert _summary_value(workbook, "Brutos sem normalização") == 0
    assert _summary_value(workbook, "Total oportunidades") == 2


def test_raw_without_listing_appears_only_in_summary(db):
    run = SearchRun.objects.create(source_name="excel")
    RawListing.objects.create(search_run=run, source_name="excel", raw_title="Bruto pendente")

    workbook = build_carol_workbook()

    assert workbook["Pesquisa Carol"].max_row == 1
    assert _summary_value(workbook, "Total bruto coletado") == 1
    assert _summary_value(workbook, "Total normalizado") == 0
    assert _summary_value(workbook, "Brutos sem normalização") == 1


def test_endpoint_returns_dated_xlsx_file(client, demo_run):
    response = client.get("/api/v1/exports/carol-xlsx")

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "jarjour-pesquisa-carol-" in response["Content-Disposition"]
    assert response["Content-Disposition"].endswith('.xlsx"')

    workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == CAROL_SHEET_NAMES
