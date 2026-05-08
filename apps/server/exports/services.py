from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from market.models import ConfidenceStatus, Listing, ReviewStatus, SearchRun
from market.selectors.summary import market_summary
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CAROL_SHEET_NAMES = [
    "Pesquisa Carol",
    "Oportunidades",
    "Revisar",
    "Links",
    "Resumo da Coleta",
]

CAROL_HEADERS = [
    "ENDEREÇO",
    "TIPO",
    "M²",
    "LOCALIZAÇÃO",
    "ALUGUEL",
    "ALUGUEL R$/M²",
    "COND.",
    "COND. R$/M²",
    "CONTATO",
    "TELEFONE",
    "OBSERVAÇÕES",
    "LINK",
    "STATUS",
    "DATA DA CAPTURA",
    "FONTE",
]

LINK_HEADERS = ["FONTE", "BAIRRO", "TIPO", "LINK", "STATUS"]

SUMMARY_LABELS = [
    "Data de geração do relatório",
    "Total bruto coletado",
    "Total normalizado",
    "Brutos sem normalização",
    "Total confiável",
    "Total para revisar",
    "Total incompleto",
    "Total duplicado provável",
    "Total descartado",
    "Total oportunidades",
    "Bairros presentes",
    "Tipos de imóvel presentes",
    "Fontes presentes",
]

REVIEW_CONFIDENCE_STATUSES = [
    ConfidenceStatus.REVIEW,
    ConfidenceStatus.INCOMPLETE,
    ConfidenceStatus.PROBABLE_DUPLICATE,
]

MONEY_FORMAT = 'R$ #,##0.00'
NUMBER_FORMAT = '#,##0.00'
DATE_FORMAT = 'dd/mm/yyyy hh:mm'
HEADER_FILL = PatternFill("solid", fgColor="F3F4F6")
SUMMARY_FILL = PatternFill("solid", fgColor="E5E7EB")
THIN_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))


def carol_export_filename(generated_at: datetime | None = None) -> str:
    generated_at = generated_at or timezone.localtime()
    return f"jarjour-pesquisa-carol-{generated_at:%Y%m%d}.xlsx"


def _base_queryset(queryset=None, search_run: SearchRun | None = None):
    listings = queryset if queryset is not None else Listing.objects.all()
    listings = listings.select_related("raw_listing")
    if search_run is not None:
        listings = listings.filter(raw_listing__search_run=search_run)
    return listings


def _non_discarded_listings(queryset=None, search_run: SearchRun | None = None):
    return _base_queryset(queryset, search_run).exclude(
        confidence_status=ConfidenceStatus.DISCARDED
    )


def _excel_datetime(value: datetime) -> datetime:
    return timezone.localtime(value).replace(tzinfo=None)


def _capture_date(listing: Listing):
    if listing.raw_listing_id and listing.raw_listing.captured_at:
        return _excel_datetime(listing.raw_listing.captured_at)
    return _excel_datetime(listing.created_at)


def _listing_row(listing: Listing) -> list:
    return [
        listing.endereco,
        listing.tipo_imovel,
        listing.area_m2,
        listing.bairro,
        listing.aluguel,
        listing.aluguel_m2,
        listing.condominio,
        listing.condominio_m2,
        listing.contato,
        listing.telefone,
        listing.observacoes,
        listing.source_url,
        listing.get_confidence_status_display(),
        _capture_date(listing),
        listing.source_name,
    ]


def _append_listing_sheet(ws: Worksheet, listings) -> None:
    ws.append(CAROL_HEADERS)
    for listing in listings:
        ws.append(_listing_row(listing))
        link_cell = ws.cell(row=ws.max_row, column=12)
        if listing.source_url:
            link_cell.hyperlink = listing.source_url
            link_cell.style = "Hyperlink"
    _style_table_sheet(ws)
    _format_listing_columns(ws)


def _append_links_sheet(ws: Worksheet, listings) -> None:
    ws.append(LINK_HEADERS)
    for listing in listings:
        ws.append(
            [
                listing.source_name,
                listing.bairro,
                listing.tipo_imovel,
                listing.source_url,
                listing.get_confidence_status_display(),
            ]
        )
        link_cell = ws.cell(row=ws.max_row, column=4)
        if listing.source_url:
            link_cell.hyperlink = listing.source_url
            link_cell.style = "Hyperlink"
    _style_table_sheet(ws)
    _set_widths(ws, {1: 18, 2: 18, 3: 20, 4: 48, 5: 24})


def _append_summary_sheet(
    ws: Worksheet,
    *,
    listings,
    search_run: SearchRun | None,
    generated_at: datetime,
) -> None:
    summary = market_summary()
    scoped_listings = list(listings)
    bairros = sorted({listing.bairro for listing in scoped_listings if listing.bairro})
    tipos = sorted({listing.tipo_imovel for listing in scoped_listings if listing.tipo_imovel})
    fontes = sorted({listing.source_name for listing in scoped_listings if listing.source_name})

    ws.append(["Resumo da Coleta", ""])
    ws.append(["Métrica", "Valor"])
    rows = [
        ("Data de geração do relatório", _excel_datetime(generated_at)),
        ("Total bruto coletado", summary["total_raw_collected"]),
        ("Total normalizado", summary["total_normalized"]),
        ("Brutos sem normalização", summary["total_raw_without_listing"]),
        ("Total confiável", summary["total_reliable"]),
        ("Total para revisar", summary["total_review"]),
        ("Total incompleto", summary["total_incomplete"]),
        ("Total duplicado provável", summary["total_probable_duplicates"]),
        ("Total descartado", summary["total_discarded"]),
        ("Total oportunidades", summary["total_opportunities"]),
        ("Bairros presentes", ", ".join(bairros) or "-"),
        ("Tipos de imóvel presentes", ", ".join(tipos) or "-"),
        ("Fontes presentes", ", ".join(fontes) or "-"),
    ]
    if search_run is not None:
        rows.append(("Recorte da coleta", str(search_run)))
    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws["A1"].font = Font(bold=True, size=14, color="111827")
    ws["A1"].fill = SUMMARY_FILL
    ws["B1"].fill = SUMMARY_FILL
    for cell in ws[2]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
    for cell in ws["B"]:
        if isinstance(cell.value, datetime):
            cell.number_format = DATE_FORMAT


def _style_table_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="111827")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column == 11)
            cell.border = THIN_BORDER


def _format_listing_columns(ws: Worksheet) -> None:
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
        for column in [5, 6, 7, 8]:
            value = ws.cell(row=row, column=column).value
            if isinstance(value, Decimal):
                ws.cell(row=row, column=column).value = float(value)
            ws.cell(row=row, column=column).number_format = MONEY_FORMAT
        ws.cell(row=row, column=14).number_format = DATE_FORMAT
    _set_widths(
        ws,
        {
            1: 28,
            2: 20,
            3: 10,
            4: 18,
            5: 14,
            6: 16,
            7: 14,
            8: 16,
            9: 24,
            10: 18,
            11: 36,
            12: 42,
            13: 22,
            14: 20,
            15: 18,
        },
    )


def _set_widths(ws: Worksheet, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def build_carol_workbook(queryset=None, search_run: SearchRun | None = None) -> Workbook:
    wb = Workbook()
    generated_at = timezone.localtime()
    all_listings = _base_queryset(queryset, search_run)
    active_listings = _non_discarded_listings(queryset, search_run)

    pesquisa = wb.active
    pesquisa.title = "Pesquisa Carol"
    _append_listing_sheet(pesquisa, active_listings)

    oportunidades = wb.create_sheet("Oportunidades")
    _append_listing_sheet(oportunidades, active_listings.filter(is_opportunity=True))

    revisar = wb.create_sheet("Revisar")
    _append_listing_sheet(
        revisar,
        active_listings.filter(
            confidence_status__in=REVIEW_CONFIDENCE_STATUSES,
        )
        | active_listings.filter(review_status=ReviewStatus.NEEDS_REVIEW),
    )

    links = wb.create_sheet("Links")
    _append_links_sheet(links, all_listings)

    resumo = wb.create_sheet("Resumo da Coleta")
    _append_summary_sheet(
        resumo,
        listings=all_listings,
        search_run=search_run,
        generated_at=generated_at,
    )

    return wb


def build_carol_workbook_response() -> HttpResponse:
    wb = build_carol_workbook()
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = carol_export_filename()
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def save_carol_workbook(path: Path | None = None) -> Path:
    export_dir = settings.JARJOUR_EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)
    path = path or export_dir / carol_export_filename()
    build_carol_workbook().save(path)
    return path


def export_carol_xlsx_view(request):
    return build_carol_workbook_response()
